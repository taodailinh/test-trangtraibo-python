import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from datetime import datetime

danhSachTrai = ["IA PÚCH", "EA H'LEO", "KOUN MOM", "SNUOL", "SAYSETTHA"]

trai = 4

# Add this to keep webdriver stay running
os.environ["PATH"] += "C:/Users/taoda/test/selenium/env"
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=options)

# maximize the window
driver.maximize_window()


# Đăng nhập
driver.get("https://dev-trangtraibo.aristqnu.com/")

driver.implicitly_wait(10)

username = driver.find_element(By.ID, "Input_UserName")

username.send_keys("chinhntk")


password = driver.find_element(By.ID, "password-field")

password.send_keys("thagrico")

form = driver.find_element(By.CLASS_NAME, "signin-form")
form.submit()


chonTrai = Select(
    driver.find_element(By.XPATH, "/html/body/div[1]/app/nav/ul[1]/select")
)
chonTrai.select_by_visible_text(danhSachTrai[trai - 1])


wb = openpyxl.load_workbook("link.xlsx")
ws = wb.active

errorLog = openpyxl.Workbook()
errorWs = errorLog.active
currentRowErrorLog = 0

for row in ws.iter_rows(min_row=2, min_col=1, max_row=98, max_col=trai + 1):
    # for row in ws.iter_rows(min_row=2, min_col=1, max_row=9, max_col=trai + 1):
    """
    for cell in row:
        link = cell.cell(1, )
        time.sleep(3)
    """
    link = row[0].value
    capQuyen = row[trai].value
    driver.get("https://" + link)
    try:
        driver.find_element(By.XPATH, "/html/body/div[1]/app/div/div/h4")
        if capQuyen == "X":
            currentRowErrorLog += 1
            errorWs.cell(currentRowErrorLog, 1).value = (
                "failed at trai "
                + str(trai)
                + "o link"
                + link
                + ", cap quyen nhung khong vao duoc"
            )
            print(
                "failed at trai "
                + str(trai)
                + "o link"
                + link
                + ", Cap quyen nhung khong vao duoc"
            )
        else:
            pass
            currentRowErrorLog += 1
            errorWs.cell(currentRowErrorLog, 1).value = (
                "passed at trai " + str(trai) + "o link" + link + ", OK"
            )
    except NoSuchElementException:
        if capQuyen == None:
            pass
            currentRowErrorLog += 1
            errorWs.cell(currentRowErrorLog, 1).value = (
                "passed at trai " + str(trai) + "o link" + link + ", OK"
            )

        else:
            print(
                "failed at trai "
                + str(trai)
                + "o link"
                + link
                + ", Khong cap quyen nhung vao duoc"
            )
            currentRowErrorLog += 1
            errorWs.cell(currentRowErrorLog, 1).value = (
                "failed at trai "
                + str(trai)
                + "o link"
                + link
                + ", Khong cap quyen nhung vao duoc"
            )
    time.sleep(3)
fileName = datetime.now().strftime("%Y%B%d%H%M%S.xlsx")
errorLog.save(fileName)
driver.quit()
""""
# Connect to db
client = MongoClient("mongodb://thagrico:Abc%40%23%24123321@45.119.84.161:27017/")
db = client["quanlytrangtrai_2807_clone"]
boNhapTrai = db["BoNhapTrai"]


# Add this to keep webdriver stay running
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=options)

# maximize the window
driver.maximize_window()

driver.get("https://dev-trangtraibo.aristqnu.com/")



driver.implicitly_wait(10)

"""
