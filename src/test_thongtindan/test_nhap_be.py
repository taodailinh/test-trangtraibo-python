from pymongo import MongoClient
from openpyxl import Workbook
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time
from selenium.webdriver import Keys, ActionChains


os.environ["PATH"] += "C:/Users/taoda/test/selenium/env"


def nhapbe(client: MongoClient, dbName, page, collectionName="BoNhapTrai"):
    # Kết nối db
    # client = MongoClient("mongodb://localhost:27017/")
    database = client[dbName]
    boNhapTrai = database[collectionName]
    # db = client["quanlytrangtrai_0910"]
    # Tìm 1 con bò bằng sô tai và in số chip
    # for i in boNhapTrai.find({"SoTai": "F040923"}):
    #     print(i["SoChip"])

    # Đém tất cả bê bò trong danh sách bò

    # Add this to keep webdriver stay running
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)

    driver = webdriver.Chrome(options=options)

    # maximize the window
    driver.maximize_window()

    driver.get(page)

    driver.implicitly_wait(10)

    username = driver.find_element(By.ID, "Input_UserName")

    username.send_keys("admin")

    password = driver.find_element(By.ID, "password-field")

    password.send_keys("admintest")

    form = driver.find_element(By.CLASS_NAME, "signin-form")
    form.submit()

    driver.implicitly_wait(10)

    driver.get(page + "quanlydan/danhsachdan")

    time.sleep(2)
    """
    danhSachDan = countCow(driver, "e-pagecountmsg")

    print(danhSachDan)
    ws = wb.active
    ws["A1"] = "Bê bò"
    ws["B1"] = danhSachDan
    if danhSachDan == 0:
        ws["C1"] = "Load chậm hơn 10s"
    if danhSachDan != soBo:
        ws["D1"] = "Không khớp query từ db"
    # danhSachDan = cowCounter.text[cowCounter.text.find("(") + 1 : cowCounter.text.find(" ")]

    # print(danhSachDan)
    """

    # Nhập bê
    wb = openpyxl.load_workbook("benhap.xlsx")
    ws = wb.active
    nhapBe = driver.find_element(
        By.XPATH, "//*[@id='div-show-column-id']/div[2]/div/span[5]/button[1]"
    )
    nhapBe.click()
    time.sleep(2)
    # Nhập số tai
    soTaiBe = driver.find_element(
        By.XPATH,
        "/html/body/div[2]/div[1]/div[2]/div/div/div[2]/form/div/div[3]/div/input",
    )
    soTaiBe.send_keys(ws["B2"].value)

    # Nhập số ID
    nhanId = driver.find_element(By.XPATH, "//label[contains(text(),'Số ID')]")
    soID = nhanId.find_element(By.XPATH, "preceding-sibling::*")
    soID.send_keys(ws["C2"].value)

    # Màu lông
    mauLong = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Màu lông')]",
    ).find_element(By.XPATH, "..")
    mauLong.click()
    ActionChains(driver).send_keys(ws["D2"].value).send_keys(Keys.ENTER).perform()

    # Giống bê
    giongBe = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Giống bê')]",
    ).find_element(By.XPATH, "..")
    giongBe.click()
    time.sleep(2)
    ActionChains(driver).send_keys(ws["F2"].value).send_keys(Keys.ENTER).perform()
    # giongBeDuocChon = driver.find_element(By.XPATH, "//li[contains(text(),'Senepol')]")
    # giongBeDuocChon.click()
    # driver.execute_script("arguments[0].value='Senepol'", giongBe)

    gioiTinh = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Giới tính')]",
    ).find_element(By.XPATH, "..")
    gioiTinh.click()
    ActionChains(driver).send_keys(ws["G2"].value).send_keys(Keys.ENTER).perform()

    soTaiMe = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Số tai mẹ')]",
    ).find_element(By.XPATH, "preceding-sibling::*")
    soTaiMe.send_keys(ws["H2"].value)

    soIdMe = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Số ID mẹ')]",
    ).find_element(By.XPATH, "preceding-sibling::*")
    soIdMe.send_keys(ws["I2"].value)

    time.sleep(5)

    trongLuong = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Trọng lượng sơ sinh ')]",
    ).find_element(By.XPATH, "preceding-sibling::*")
    trongLuong.send_keys(ws["J2"].value)

    giongCha = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Giống cha')]",
    ).find_element(By.XPATH, "preceding-sibling::*")
    giongCha.send_keys(ws["K2"].value)

    oChuong = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Ô chuồng')]",
    ).find_element(By.XPATH, "..")
    oChuong.click()
    ActionChains(driver).send_keys(ws["L2"].value).send_keys(Keys.ENTER).perform()

    nhomBe = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Nhóm bê')]",
    ).find_element(By.XPATH, "..")
    nhomBe.click()
    ActionChains(driver).send_keys(ws["M2"].value).send_keys(Keys.ENTER).perform()

    nguoiBaoDe = driver.find_element(
        By.XPATH,
        "//label[contains(text(),'Người báo đẻ')]",
    ).find_element(By.XPATH, "..")
    nguoiBaoDe.click()
    ActionChains(driver).send_keys("D").send_keys(Keys.ENTER).perform()
    # Hoàn   tất lưu bê sinh
    driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[4]/button[2]").click()

    time.sleep(10)
    boDuocThem = boNhapTrai.find({"SoTai": "F202300609"})
    for bo in boDuocThem:
        print(bo.get("SoTai"))
    time.sleep(10)

    driver.quit()
