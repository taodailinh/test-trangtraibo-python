import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from datetime import datetime, timedelta


danhSachTrai = ["IA PÚCH", "EA H'LEO", "KOUN MOM", "SNUOL", "SAYSETTHA"]

def testXuatThongTinDan(client,drivers, page, trai=1):

    db = client["LinhTest"]
    col = db["TestOnDinh"]

    timeToStop = datetime.now()+timedelta(minutes=60)
    for driver in drivers:
        driver.get(page+"quanlydan/danhsachdan")


    while datetime.now() < timeToStop:
        for driver in drivers:
            time.sleep(5)
            print(EC.element_to_be_clickable((By.XPATH,"//button[text()='Xuất Excel']")))
            try:
                if EC.element_to_be_clickable((By.XPATH,"//button[text()='Xuất Excel']")) != False:
                    driver.find_element(By.XPATH,"//button[text()='Xuất Excel']").click()
            except Exception as e:
                print("Đã có lỗi xảy ra khi bấm click xuất excel",e)
    for driver in drivers:
        driver.quit()
"""    
    try:
        try:
            for driver in drivers:
                wait = WebDriverWait(driver,300)
                # driver.implicitly_wait(20)
                xuatexcel = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/app/div/div/div/div[2]/div/span[5]/button[3]")))
                xuatexcel.click()
                time.sleep(30)
        except Exception as e:
            print("Đã có lỗi khi click xuất báo cáo lần đầu: ",e)
        for i in range(2,10):
            try:
                for driver in drivers:
                    # time.sleep(2)
                    wait = WebDriverWait(driver,300)
                    spinner = wait.until(EC.presence_of_element_located((By.XPATH,"/html/body/div[1]/app/div/div/div/div[5]/div[1]")))
                    # spinner = driver.find_element(By.XPATH,"/html/body/div[1]/app/div/div/div/div[5]/div[1]")
                    # wait.until(EC.staleness_of(spinner))
                    wait.until(EC.invisibility_of_element(spinner))
                    time.sleep(5)
                    print("done")
                    driver.get(page+"quanlydan/danhsachdan")
                    xuatexcel = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/app/div/div/div/div[2]/div/span[5]/button[3]")))
                    xuatexcel.click()
                    time.sleep(30)
            except Exception as e:
                print("Đã có lỗi khi click xuất báo cáo lần "+str(i),e)
        try:
            for driver in drivers:
                driver.quit()
        except Exception as e:
            print("Đã có lỗi khi thoát driver: ",e)
    except Exception as e:
        print("Đã có lỗi xảy ra:+",e)
        for driver in drivers:
            driver.quit()
"""

    # Đăng nhập

    # chonTrai = Select(
    #     driver.find_element(By.XPATH, "/html/body/div[1]/app/nav/ul[1]/select")
    # )
    # chonTrai.select_by_visible_text(danhSachTrai[trai - 1])
    # xuatexcel = driver.find_element(By.XPATH,"//button[contains(text(), 'Xuất Excel')]")

    # script_path = os.path.dirname(os.path.abspath(__file__))

    # workbook_path = os.path.join(script_path, "link.xlsx")

    # wb = openpyxl.load_workbook(workbook_path)
    # ws = wb.active



"""

    for row in ws.iter_rows(min_row=2, min_col=1, max_row=98, max_col=trai + 1):
        # for row in ws.iter_rows(min_row=2, min_col=1, max_row=9, max_col=trai + 1):
        for cell in row:
            link = cell.cell(1, )
            time.sleep(3)
        link = row[0].value
        capQuyen = row[trai].value
        print(capQuyen)
        driver.get(page + link)
        result = ""
        try:
            div = driver.find_element(By.XPATH,"//h4[contains(text(), 'Bạn không có quyền')]")
            print(div)
            # driver.find_element(By.XPATH, "/html/body/div[1]/app/div/div/h4")
            if capQuyen == "x":
                result = (
                    "failed at trai "
                    + str(trai)
                    + " o link "
                    + link
                    + ", cap quyen nhung khong vao duoc"
                )
                excelWriter.append([result])
                print(result)
            else:
                result = (
                    "passed at trai " + str(trai) + " o link " + link + ", OK. Không cấp quyền và không vào được"
                )                
                excelWriter.append([result])
                print(result)
        except NoSuchElementException:
            if capQuyen == "x":
                result = (
                    "passed at trai " + str(trai) + " o link " + link + ", OK. Cấp quyền và vào được"
                )
                excelWriter.append([result])
                print(result)
            else:
                result = (
                    "failed at trai "
                    + str(trai)
                    + " o link "
                    + link
                    + ", Khong cap quyen nhung vao duoc"
                )
                excelWriter.append([result])
                print(result)
        time.sleep(3)
        """
