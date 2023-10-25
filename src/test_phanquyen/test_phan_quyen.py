import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from datetime import datetime


def testPhanQuyenUser(user, pw, page,excelWriter, trai=1):
    danhSachTrai = ["IA PÚCH", "EA H'LEO", "KOUN MOM", "SNUOL", "SAYSETTHA"]

    # Add this to keep webdriver stay running
    os.environ["PATH"] += "C:/Users/taoda/test/selenium/env"
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)

    driver = webdriver.Chrome(options=options)

    # maximize the window
    driver.maximize_window()

    # Đăng nhập
    driver.get(page)

    driver.implicitly_wait(3)

    username = driver.find_element(By.ID, "Input_UserName")

    username.send_keys(str(user))

    password = driver.find_element(By.ID, "password-field")

    password.send_keys(str(pw))

    form = driver.find_element(By.CLASS_NAME, "signin-form")
    form.submit()

    chonTrai = Select(
        driver.find_element(By.XPATH, "/html/body/div[1]/app/nav/ul[1]/select")
    )
    chonTrai.select_by_visible_text(danhSachTrai[trai - 1])

    script_path = os.path.dirname(os.path.abspath(__file__))
    # print("Current path")
    # print(script_path)

    workbook_path = os.path.join(script_path, "link.xlsx")

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, min_col=1, max_row=98, max_col=trai + 1):
        # for row in ws.iter_rows(min_row=2, min_col=1, max_row=9, max_col=trai + 1):
        """
        for cell in row:
            link = cell.cell(1, )
            time.sleep(3)
        """
        link = row[0].value
        capQuyen = row[trai].value
        print(capQuyen)
        driver.get(page + link)
        result = ""
        try:
            div = driver.find_element(By.XPATH,"//h4[contains(text(), 'Bạn không có quyền')]")
            print(div)
            # driver.find_element(By.XPATH, "/html/body/div[1]/app/div/div/h4")
            if capQuyen == "x":
                result = (
                    "failed at trai "
                    + str(trai)
                    + " o link "
                    + link
                    + ", cap quyen nhung khong vao duoc"
                )
                excelWriter.append([result])
                print(result)
            else:
                result = (
                    "passed at trai " + str(trai) + " o link " + link + ", OK. Không cấp quyền và không vào được"
                )                
                excelWriter.append([result])
                print(result)
        except NoSuchElementException:
            if capQuyen == "x":
                result = (
                    "passed at trai " + str(trai) + " o link " + link + ", OK. Cấp quyền và vào được"
                )
                excelWriter.append([result])
                print(result)
            else:
                result = (
                    "failed at trai "
                    + str(trai)
                    + " o link "
                    + link
                    + ", Khong cap quyen nhung vao duoc"
                )
                excelWriter.append([result])
                print(result)
        time.sleep(3)
    driver.quit()
    """"
    # Connect to db
    client = MongoClient("mongodb://thagrico:Abc%40%23%24123321@45.119.84.161:27017/")
    db = client["quanlytrangtrai_2807_clone"]
    boNhapTrai = db["BoNhapTrai"]


    # Add this to keep webdriver stay running
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)

    driver = webdriver.Chrome(options=options)

    # maximize the window
    driver.maximize_window()

    driver.get("https://dev-trangtraibo.aristqnu.com/")



    driver.implicitly_wait(10)

    """
