from pymongo import MongoClient
from openpyxl import Workbook
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time
from selenium.webdriver import Keys, ActionChains

from countCount import countCow
from query import cowCount

os.environ["PATH"] += "C:/Users/taoda/test/selenium/env"


# Kết nối db
client = MongoClient("mongodb://thagrico:Abc%40%23%24123321@45.119.84.161:27017/")
db = client["quanlytrangtrai_2807_clone"]
boNhapTrai = db["BoNhapTrai"]


# Tìm 1 con bò bằng sô tai và in số chip
# for i in boNhapTrai.find({"SoTai": "F040923"}):
#     print(i["SoChip"])

# Đém tất cả bê bò trong danh sách bò
soBo = cowCount(boNhapTrai, "NhomBo", "Bo") + cowCount(boNhapTrai, "NhomBo", "Be")
print(soBo)

# Add this to keep webdriver stay running
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=options)

# maximize the window
driver.maximize_window()

driver.get("https://test-trangtraibo.aristqnu.com/")

driver.implicitly_wait(10)

username = driver.find_element(By.ID, "Input_UserName")

username.send_keys("admin")


password = driver.find_element(By.ID, "password-field")

password.send_keys("admintest")

form = driver.find_element(By.CLASS_NAME, "signin-form")
form.submit()

driver.implicitly_wait(10)

driver.get("https://test-trangtraibo.aristqnu.com/quanlydan/danhsachdan")

wb = Workbook()


time.sleep(10)

danhSachDan = countCow(driver, "e-pagecountmsg")

print(danhSachDan)
ws = wb.active
ws["A1"] = "Bê bò"
ws["B1"] = danhSachDan
if danhSachDan == 0:
    ws["C1"] = "Load chậm hơn 10s"
if danhSachDan != soBo:
    ws["D1"] = "Không khớp query từ db"
# danhSachDan = cowCounter.text[cowCounter.text.find("(") + 1 : cowCounter.text.find(" ")]

# print(danhSachDan)


# Nhập bê
nhapBe = driver.find_element(
    By.XPATH, "//*[@id='div-show-column-id']/div[2]/div/span[5]/button[1]"
)
nhapBe.click()
time.sleep(2)
# Nhập số tai
soTaiBe = driver.find_element(
    By.XPATH,
    "/html/body/div[2]/div[1]/div[2]/div/div/div[2]/form/div/div[3]/div/input",
)
soTaiBe.send_keys("F202300609")

# Nhập số ID
nhanId = driver.find_element(By.XPATH, "//label[contains(text(),'Số ID')]")
print(nhanId.text)
soID = nhanId.find_element(By.XPATH, "preceding-sibling::*")
soID.send_keys("F202300609ID")

# Màu lông
mauLong = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Màu lông')]",
).find_element(By.XPATH, "preceding-sibling::*")
driver.execute_script("arguments[0].value='Nâu'", mauLong)

# Giống bê
giongBe = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Giống bê')]",
).find_element(By.XPATH, "..")
giongBe.click()
time.sleep(2)
ActionChains(driver).send_keys("Senepol").send_keys(Keys.ENTER).perform()
# giongBeDuocChon = driver.find_element(By.XPATH, "//li[contains(text(),'Senepol')]")
# giongBeDuocChon.click()
# driver.execute_script("arguments[0].value='Senepol'", giongBe)

gioiTinh = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Giới tính')]",
).find_element(By.XPATH, "preceding-sibling::*")
driver.execute_script("arguments[0].value='Cái'", gioiTinh)

soTaiMe = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Số tai mẹ')]",
).find_element(By.XPATH, "preceding-sibling::*")
soTaiMe.send_keys("I0920B01129")

soIdMe = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Số ID mẹ')]",
).find_element(By.XPATH, "preceding-sibling::*")
soIdMe.send_keys("982 123708768054")

time.sleep(3)

trongLuong = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Trọng lượng sơ sinh ')]",
).find_element(By.XPATH, "preceding-sibling::*")
trongLuong.send_keys("30")

giongCha = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Giống cha')]",
).find_element(By.XPATH, "preceding-sibling::*")
trongLuong.send_keys("Droughtmaster")

oChuong = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Ô chuồng')]",
).find_element(By.XPATH, "preceding-sibling::*")
driver.execute_script("arguments[0].value='A2.02'", oChuong)


nhomBe = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Nhóm bê')]",
).find_element(By.XPATH, "..")
nhomBe.click()
ActionChains(driver).send_keys("Nhóm bê").send_keys(Keys.ENTER).perform()


nguoiBaoDe = driver.find_element(
    By.XPATH,
    "//label[contains(text(),'Người báo đẻ')]",
).find_element(By.XPATH, "preceding-sibling::*")
driver.execute_script("arguments[0].value='Ksor Minh'", nguoiBaoDe)

# Hoàn tất lưu bê sinh
driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[4]/button[2]").click()


time.sleep(10)
driver.switch_to.new_window("tab")
driver.get("https://test-trangtraibo.aristqnu.com/thuy/danhsachthuy")


# Kiểm tra tab thú y có load được danh sách bò nhâp viện không, điền file excel số lượng bò đang nằm viện
time.sleep(10)
boNamVien = countCow(driver, "e-pagecountmsg")

print(boNamVien)

ws["A2"] = "Bò nằm viện"
ws["B2"] = boNamVien
if boNamVien == 0:
    ws["C2"] = "Load chậm hơn 10s"

wb.save("20230904.xlsx")
# driver.quit()
